from pathlib import Path

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from api.routes.output import router as output_router
from api.routes.sessions import get_manager
from api.routes.sessions import router as sessions_router
from api.state import SessionManager
from core.db.connection import close_all, open_connection
from core.db.migrations import init_schema


def _scan_result(per_file: dict):
    """ScanResult de filename_glob con per_file poblado."""
    from core.scanners.base import ConfidenceLevel, ScanResult

    return ScanResult(
        count=sum(per_file.values()),
        confidence=ConfidenceLevel.HIGH,
        method="filename_glob",
        breakdown={},
        flags=[],
        errors=[],
        files_scanned=len(per_file),
        duration_ms=10,
        per_file=per_file,
    )


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("INFORME_MENSUAL_ROOT", "A:/informe mensual")
    monkeypatch.setenv("OVERSEER_OUTPUT_DIR", str(tmp_path / "outputs"))
    app = FastAPI()
    conn = open_connection(tmp_path / "out.db")
    init_schema(conn)
    app.dependency_overrides[get_manager] = lambda: SessionManager(conn=conn)
    app.include_router(sessions_router, prefix="/api")
    app.include_router(output_router, prefix="/api")
    yield TestClient(app)
    close_all()


def test_generate_output_creates_xlsx(client, tmp_path):
    client.post("/api/sessions", json={"year": 2026, "month": 4})
    client.post("/api/sessions/2026-04/scan", json={"scope": "all"})
    response = client.post("/api/sessions/2026-04/output", json={})
    assert response.status_code == 200
    data = response.json()
    assert Path(data["output_path"]).exists()
    assert data["output_path"].endswith(".xlsx")


def test_output_uses_v2_priority(client, tmp_path):
    """V2 cells with user_override get the override in the Excel."""
    import openpyxl

    client.post("/api/sessions", json={"year": 2026, "month": 4})
    client.post("/api/sessions/2026-04/scan", json={"scope": "all"})

    # Override HRB/odi to 17 via SessionManager
    mgr = client.app.dependency_overrides[get_manager]()
    mgr.apply_user_override("2026-04", "HRB", "odi", value=17, note="manual")

    out = client.post("/api/sessions/2026-04/output", json={}).json()
    wb = openpyxl.load_workbook(out["output_path"])
    dest = list(wb.defined_names["HRB_odi_count"].destinations)[0]
    sheet, coord = dest
    assert wb[sheet][coord].value == 17


def test_output_emits_worker_totals(client, tmp_path, monkeypatch):
    import openpyxl

    # Use tmp_path root so the charla folder doesn't contain real PDFs →
    # present_files=None (legacy: sum-all), worker_count == sum of synthetic marks.
    (tmp_path / "ABRIL").mkdir()
    monkeypatch.setenv("INFORME_MENSUAL_ROOT", str(tmp_path))
    client.post("/api/sessions", json={"year": 2026, "month": 4})
    mgr = client.app.dependency_overrides[get_manager]()
    mgr.apply_worker_count(
        "2026-04",
        "HLL",
        "charla",
        marks={"c1.pdf": [{"page": 1, "count": 18}, {"page": 2, "count": 22}]},
        status="terminado",
    )
    out = client.post("/api/sessions/2026-04/output", json={}).json()
    wb = openpyxl.load_workbook(out["output_path"])
    sheet, coord = list(wb.defined_names["HLL_workers_chgen"].destinations)[0]
    assert wb[sheet][coord].value == 40

    # una celda nunca contada no se emite: su rango con nombre queda en blanco
    nc_sheet, nc_coord = list(wb.defined_names["HLL_workers_chintegral"].destinations)[0]
    assert wb[nc_sheet][nc_coord].value is None


def test_worker_warnings_flag_incomplete_cell(client, tmp_path):
    client.post("/api/sessions", json={"year": 2026, "month": 4})
    mgr = client.app.dependency_overrides[get_manager]()
    mgr.apply_filename_result("2026-04", "HLL", "charla", _scan_result({"c1.pdf": 3}))
    # per_file poblado, sin worker_status → celda incompleta
    out = client.post("/api/sessions/2026-04/output", json={}).json()
    warned = {(w["hospital"], w["sigla"]) for w in out["worker_warnings"]}
    assert ("HLL", "charla") in warned


def test_worker_warnings_silent_when_terminado(client, tmp_path):
    client.post("/api/sessions", json={"year": 2026, "month": 4})
    mgr = client.app.dependency_overrides[get_manager]()
    mgr.apply_filename_result("2026-04", "HLL", "charla", _scan_result({"c1.pdf": 3}))
    mgr.apply_worker_count("2026-04", "HLL", "charla", status="terminado")
    out = client.post("/api/sessions/2026-04/output", json={}).json()
    warned = {(w["hospital"], w["sigla"]) for w in out["worker_warnings"]}
    assert ("HLL", "charla") not in warned


def test_worker_warnings_flag_en_progreso_cell(client, tmp_path):
    client.post("/api/sessions", json={"year": 2026, "month": 4})
    mgr = client.app.dependency_overrides[get_manager]()
    mgr.apply_filename_result("2026-04", "HLL", "charla", _scan_result({"c1.pdf": 3}))
    mgr.apply_worker_count("2026-04", "HLL", "charla", status="en_progreso")
    out = client.post("/api/sessions/2026-04/output", json={}).json()
    warned = {(w["hospital"], w["sigla"]) for w in out["worker_warnings"]}
    assert ("HLL", "charla") in warned


def test_build_report_title_uses_session_month():
    """The report title must reflect the session's month/year, not the template's
    hardcoded 'MARZO 2026'. The project line stays constant."""
    from api.routes.output import _build_report_title

    assert _build_report_title(2026, 5) == (
        "RESUMEN EJECUTIVO ACTIVIDADES DE PREVENCIÓN DE RIESGOS\n"
        "MAYO 2026\n"
        "PROYECTO RED LOS RÍOS - LOS LAGOS"
    )
    assert "ABRIL 2026" in _build_report_title(2026, 4)
    assert "DICIEMBRE 2027" in _build_report_title(2027, 12)


def test_output_writes_dynamic_month_title(client, tmp_path):
    """End-to-end: the generated RESUMEN's title cell shows the session month."""
    import openpyxl

    client.post("/api/sessions", json={"year": 2026, "month": 4})
    out = client.post("/api/sessions/2026-04/output", json={}).json()
    wb = openpyxl.load_workbook(out["output_path"])
    sheet, coord = list(wb.defined_names["report_title"].destinations)[0]
    val = str(wb[sheet][coord.replace("$", "")].value)
    assert "ABRIL 2026" in val
    assert "MARZO" not in val


def test_build_cell_values_emits_zero_for_uncounted_cells():
    """Cells absent from session state (a hospital not yet counted) must be written
    as 0, not left blank. Regression for the 2026-06-06 empty-cell report."""
    from api.routes.output import _build_cell_values
    from core.domain import HOSPITALS, SIGLAS

    state = {"cells": {"HLL": {"art": {"user_override": 5}}}}
    vals = _build_cell_values(state)
    assert vals["HLL_art_count"] == 5
    assert vals["HLU_reunion_count"] == 0  # never counted → explicit 0
    assert len(vals) == len(HOSPITALS) * len(SIGLAS)


def test_build_cell_values_skips_excluded():
    from api.routes.output import _build_cell_values
    from core.domain import HOSPITALS, SIGLAS

    state = {"cells": {"HLL": {"art": {"user_override": 5, "excluded": True}}}}
    vals = _build_cell_values(state)
    assert "HLL_art_count" not in vals  # excluded → skipped, left blank
    assert len(vals) == len(HOSPITALS) * len(SIGLAS) - 1


def test_build_cell_values_honors_per_file_overrides():
    """End-to-end of the writer fix at the route layer: a per-file-corrected cell
    emits the corrected count, not the stale filename count."""
    from api.routes.output import _build_cell_values

    state = {
        "cells": {
            "HLL": {
                "charla": {
                    "filename_count": 1,
                    "per_file": {"comp.pdf": 1},
                    "per_file_overrides": {"comp.pdf": 486},
                }
            }
        }
    }
    assert _build_cell_values(state)["HLL_charla_count"] == 486


def test_worker_warnings_silent_when_no_pdfs(client, tmp_path):
    client.post("/api/sessions", json={"year": 2026, "month": 4})
    mgr = client.app.dependency_overrides[get_manager]()
    # worker_status presente pero sin per_file → la celda no tiene nada que contar
    mgr.apply_worker_count("2026-04", "HLL", "charla", status="en_progreso")
    out = client.post("/api/sessions/2026-04/output", json={}).json()
    warned = {(w["hospital"], w["sigla"]) for w in out["worker_warnings"]}
    assert ("HLL", "charla") not in warned


# ---------------------------------------------------------------------------
# Task 3: DIFPTS_WORKER_HOSPITALS + emit HPV_workers_difpts
# ---------------------------------------------------------------------------


def test_build_worker_values_emits_difpts_for_hpv(tmp_path):
    from api.routes.output import _build_worker_values

    state = {
        "month_root": str(tmp_path / "nope"),  # non-existent → present=None → sum-all marks
        "cells": {
            "HPV": {
                "dif_pts": {
                    "worker_marks": {"d1.pdf": [{"page": 1, "count": 12}, {"page": 2, "count": 8}]}
                }
            },
        },
    }
    assert _build_worker_values(state)["HPV_workers_difpts"] == 20


def test_build_worker_values_difpts_zero_when_uncounted(tmp_path):
    from api.routes.output import _build_worker_values

    state = {
        "month_root": str(tmp_path / "nope"),
        "cells": {"HPV": {"dif_pts": {"per_file": {"d1.pdf": 1}}}},  # no worker_marks
    }
    assert _build_worker_values(state)["HPV_workers_difpts"] == 0


def test_build_worker_values_difpts_not_emitted_for_non_hpv(tmp_path):
    from api.routes.output import _build_worker_values

    state = {
        "month_root": str(tmp_path / "nope"),
        "cells": {"HRB": {"dif_pts": {"worker_marks": {"d1.pdf": [{"page": 1, "count": 99}]}}}},
    }
    vals = _build_worker_values(state)
    assert "HRB_workers_difpts" not in vals
    assert "HPV_workers_difpts" not in vals  # HPV has no dif_pts cell here
