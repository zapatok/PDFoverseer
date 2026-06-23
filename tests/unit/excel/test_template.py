from pathlib import Path

import pytest

from core.excel.template import DEFAULT_TEMPLATE, list_named_ranges, load_template


def test_default_template_exists():
    assert DEFAULT_TEMPLATE.exists()


def test_load_template_returns_workbook():
    wb = load_template(DEFAULT_TEMPLATE)
    assert any("Cump" in s for s in wb.sheetnames)


def test_list_named_ranges_includes_expected_count_cells():
    wb = load_template(DEFAULT_TEMPLATE)
    names = list_named_ranges(wb)
    # 76 cantidad cells: 4 hospitals × 19 written siglas (20 − chps, excluded from
    # the Excel in Incr B); revdocmaq/espacios wired to the former orphan rows.
    cantidad_names = [n for n in names if n.endswith("_count")]
    assert len(cantidad_names) == 76
    assert "HPV_art_count" in names
    assert "HLL_reunion_count" in names
    assert "HRB_revdocmaq_count" in names
    assert "HLL_espacios_count" in names
    assert "HRB_chps_count" not in names  # chps no longer written to the Excel


def test_resumen_writes_new_categories_to_orphan_rows(tmp_path):
    """Incr B end-to-end: the writer fills the former orphan rows B22/B26 from the
    new named ranges (HRB→K, HLL→G; revdocmaq=row 22, espacios=row 26)."""
    import openpyxl

    from core.excel.writer import generate_resumen

    out = tmp_path / "r.xlsx"
    generate_resumen(
        cell_values={"HRB_revdocmaq_count": 3, "HLL_espacios_count": 2, "report_title": "T"},
        output_path=out,
    )
    ws = openpyxl.load_workbook(out).active
    assert ws["K22"].value == 3  # HRB revdocmaq → row 22
    assert ws["G26"].value == 2  # HLL espacios → row 26
