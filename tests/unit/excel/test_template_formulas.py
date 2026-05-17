"""El template debe traer las fórmulas de HH correctas y las celdas de
trabajadores en blanco (spec §8.2)."""

import openpyxl
import pytest

from core.excel.template import DEFAULT_TEMPLATE

HH_COLS = ("H", "J", "L", "N")


@pytest.fixture(scope="module")
def ws():
    # data_only=False (por defecto) → las fórmulas se leen como texto
    return openpyxl.load_workbook(DEFAULT_TEMPLATE).active


@pytest.mark.parametrize("col", HH_COLS)
def test_hh_chgen_formula_points_to_row_29(ws, col):
    assert ws[f"{col}13"].value == f"={col}29*0.25"


@pytest.mark.parametrize("col", HH_COLS)
def test_hh_chintegral_formula_points_to_row_30(ws, col):
    assert ws[f"{col}14"].value == f"={col}30*0.5"


@pytest.mark.parametrize("col", HH_COLS)
@pytest.mark.parametrize("row", (29, 30))
def test_worker_value_cells_ship_blank(ws, col, row):
    assert ws[f"{col}{row}"].value is None
