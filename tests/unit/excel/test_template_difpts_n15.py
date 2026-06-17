"""Incr 3B: the template wires HPV dif_pts worker total to N15 (no =M15*0.5 fallback)."""

from pathlib import Path

import openpyxl

TEMPLATE = Path("data/templates/RESUMEN_template_v1.xlsx")


def _load():
    return openpyxl.load_workbook(TEMPLATE)


def test_hpv_workers_difpts_named_range_points_at_n15():
    wb = _load()
    assert "HPV_workers_difpts" in wb.defined_names
    dest = list(wb.defined_names["HPV_workers_difpts"].destinations)
    assert dest == [(wb.active.title, "$N$15")]


def test_n15_formula_cleared_to_zero():
    wb = _load()
    n15 = wb.active["N15"].value
    assert n15 != "=M15*0.5"
    assert n15 in (0, None) or isinstance(n15, (int, float))


def test_other_hospitals_row15_hh_formula_intact():
    wb = _load()
    ws = wb.active
    assert ws["H15"].value == "=G15*0.5"
    assert ws["J15"].value == "=I15*0.5"
    assert ws["L15"].value == "=K15*0.5"


def test_existing_named_ranges_intact():
    wb = _load()
    names = set(wb.defined_names)
    for n in (
        "HLL_dif_pts_count",
        "HPV_dif_pts_count",
        "HLL_workers_chgen",
        "HPV_workers_chintegral",
        "report_title",
    ):
        assert n in names
    worker_ranges = sorted(n for n in names if "_workers_" in n)
    assert len(worker_ranges) == 9  # 8 chgen/chintegral + HPV_workers_difpts


def test_logo_image_survived_round_trip():
    wb = _load()
    # The constructora logo (B2) must survive any openpyxl re-save.
    assert len(wb.active._images) >= 1
