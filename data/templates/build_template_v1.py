"""Build RESUMEN_template_v1.xlsx programmatically.

Takes the production RESUMEN_ABRIL_2026.xlsx as a layout base, then:
  1. Adds 72 named ranges for cantidad cells (4 hospitals × 18 siglas)
  2. Adds 8 named ranges for workforce cells (chgen + chintegral × 4 hospitals)
  3. Blanks any pre-existing cantidad values so the template ships clean
  4. Adds a CHPS row (sigla #18) at row 31 — the sample doesn't include it

Run from project root:
    python data/templates/build_template_v1.py

Re-running is idempotent. Bump filename suffix when shipping a v2.
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.workbook.defined_name import DefinedName

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
SRC = PROJECT_ROOT / "data" / "output_sample" / "RESUMEN_ABRIL_2026.xlsx"
DST = PROJECT_ROOT / "data" / "templates" / "RESUMEN_template_v1.xlsx"
LOGO = PROJECT_ROOT / "data" / "templates" / "logoConstructoraSur.png"

# NOTE: the shipped RESUMEN_template_v1.xlsx is the AUTHORITATIVE artifact and
# carries hand-patches not reproduced here (e.g. the 2026-06-04 O11/O12 #REF!
# fix). Re-running this recipe rebuilds from the sample and would drop those —
# treat the committed .xlsx as source of truth; keep this script in sync as a
# best-effort recipe, do not blindly regenerate over the live template.

# Rows that exist in the sample but map to no canonical sigla (no named range).
# They ship as an explicit 0 in every hospital column so the report never shows
# a blank where the other hospitals show 0 (Daniel's 2026-06-06 report).
ORPHAN_ROWS: tuple[int, ...] = (22, 26)

# Sigla → row index in the worksheet.
SIGLA_ROW: dict[str, int] = {
    "reunion": 10,
    "irl": 11,
    "odi": 12,
    "charla": 13,
    "chintegral": 14,
    "dif_pts": 15,
    "art": 16,
    "insgral": 17,
    "bodega": 18,
    "maquinaria": 19,
    "ext": 20,
    "senal": 21,
    # 22 = orphan
    "exc": 23,
    "altura": 24,
    "caliente": 25,
    # 26 = orphan
    "herramientas_elec": 27,
    "andamios": 28,
    "chps": 31,  # Added in this script — sample lacks CHPS row
}

# Hospital → cantidad column letter
CANTIDAD_COL: dict[str, str] = {"HLL": "G", "HLU": "I", "HRB": "K", "HPV": "M"}

# Hospital → HH column letter (workforce values live in HH columns)
HH_COL: dict[str, str] = {"HLL": "H", "HLU": "J", "HRB": "L", "HPV": "N"}

# Workforce purpose → row index
WORKFORCE_ROW: dict[str, int] = {"chgen": 29, "chintegral": 30}


def _set_or_replace_name(wb, name: str, attr_text: str) -> None:
    """Replace if exists, else create. openpyxl 3.1+ DefinedNameDict API."""
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=attr_text)


def build() -> tuple[int, int]:
    """Build the template. Returns (cantidad_count, workforce_count)."""
    if not SRC.exists():
        raise FileNotFoundError(f"Source workbook missing: {SRC}")
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)

    wb = openpyxl.load_workbook(DST)
    ws = wb.active
    sheet_name = ws.title

    # Add CPHS label at row 31 (the sample leaves it empty). Visible label is
    # CPHS (correct acronym); the internal sigla key stays "chps".
    ws.cell(row=31, column=2, value=("CPHS — Comité Paritario de Higiene y Seguridad"))
    ws.cell(row=31, column=6, value="—")  # periodicidad column

    # B2 holds a stale #VALUE! in the sample (a missing logo reference). Clear it
    # and drop in the constructora logo, ~110x110 px, anchored in the B2:D5 box.
    ws["B2"] = None
    if LOGO.exists():
        img = XLImage(str(LOGO))
        img.width = 110
        img.height = 110
        ws.add_image(img, "B2")

    # Orphan rows ship as explicit 0 in every hospital cantidad column.
    for row in ORPHAN_ROWS:
        for col in CANTIDAD_COL.values():
            ws[f"{col}{row}"] = 0

    # Cantidad named ranges + blank pre-existing values
    cantidad_count = 0
    for sigla, row in SIGLA_ROW.items():
        for hosp, col in CANTIDAD_COL.items():
            name = f"{hosp}_{sigla}_count"
            ref = f"'{sheet_name}'!${col}${row}"
            _set_or_replace_name(wb, name, ref)
            ws[f"{col}{row}"] = None
            cantidad_count += 1

    # Workforce named ranges (values live in HH columns, see audit)
    workforce_count = 0
    for hosp, hh_col in HH_COL.items():
        for purpose, row in WORKFORCE_ROW.items():
            name = f"{hosp}_workers_{purpose}"
            ref = f"'{sheet_name}'!${hh_col}${row}"
            _set_or_replace_name(wb, name, ref)
            workforce_count += 1

    # Fix HH formula + blank stale workforce values (spec §8.2).
    # The sample's H14 points to row 29 (chgen); chintegral lives in row 30.
    ws["H14"] = "=H30*0.5"
    for hh_col in HH_COL.values():
        for row in WORKFORCE_ROW.values():
            ws[f"{hh_col}{row}"] = None

    wb.save(DST)
    return cantidad_count, workforce_count


def verify() -> None:
    """Reload + assert all expected ranges exist + cantidad cells are blank."""
    wb = openpyxl.load_workbook(DST)
    ws = wb.active

    names = list(wb.defined_names)
    count_names = sorted(n for n in names if n.endswith("_count"))
    worker_names = sorted(n for n in names if "_workers_" in n)

    if len(count_names) != 72:
        raise AssertionError(f"Expected 72 cantidad named ranges, got {len(count_names)}")
    if len(worker_names) != 8:
        raise AssertionError(f"Expected 8 worker named ranges, got {len(worker_names)}")

    # Spot-check: a few specific named ranges resolve to correct cells
    for name, expected_coord in [
        ("HPV_art_count", f"'{ws.title}'!$M$16"),
        ("HLL_reunion_count", f"'{ws.title}'!$G$10"),
        ("HRB_chps_count", f"'{ws.title}'!$K$31"),
        ("HLU_workers_chgen", f"'{ws.title}'!$J$29"),
    ]:
        actual = wb.defined_names[name].attr_text
        if actual != expected_coord:
            raise AssertionError(f"{name}: expected {expected_coord!r}, got {actual!r}")

    # Confirm cantidad cells were blanked for at least one hospital
    for col in ("G", "I", "K", "M"):
        for row in (10, 16, 28, 31):
            val = ws[f"{col}{row}"].value
            if val is not None:
                raise AssertionError(f"Cell {col}{row} should be blank in template, has {val!r}")

    # Orphan rows must ship as 0 in every hospital column (no blanks)
    for col in CANTIDAD_COL.values():
        for row in ORPHAN_ROWS:
            if ws[f"{col}{row}"].value != 0:
                raise AssertionError(
                    f"Orphan cell {col}{row} should be 0, has {ws[f'{col}{row}'].value!r}"
                )

    # The stale #VALUE! logo cell must be cleared
    if ws["B2"].value is not None:
        raise AssertionError(f"B2 should be cleared (logo lives there), has {ws['B2'].value!r}")

    # HH formulas: row 13 = chgen ×0.25, row 14 = chintegral ×0.5
    for col in HH_COL.values():
        if ws[f"{col}13"].value != f"={col}29*0.25":
            raise AssertionError(f"{col}13 HH formula wrong: {ws[f'{col}13'].value!r}")
        if ws[f"{col}14"].value != f"={col}30*0.5":
            raise AssertionError(f"{col}14 HH formula wrong: {ws[f'{col}14'].value!r}")

    # Workforce value cells must ship blank
    for col in HH_COL.values():
        for row in WORKFORCE_ROW.values():
            if ws[f"{col}{row}"].value is not None:
                raise AssertionError(f"Cell {col}{row} should be blank")


def main() -> int:
    cant, work = build()
    verify()
    print(f"Built: {DST}")
    print(f"  cantidad named ranges: {cant}")
    print(f"  workforce named ranges: {work}")
    print(f"  total: {cant + work}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
